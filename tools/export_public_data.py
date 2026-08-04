#!/usr/bin/env python3
"""Create the public Hausbar app export from locked private v7.66 sources.

The exporter is deny-by-default: only fields named in the public allowlist are
read into the public object. Private master fields and original-image metadata
are never serialized.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from check_environment import EnvironmentContractError, verify_environment

load_workbook = None
Image = None

ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "config"

ID_RE = re.compile(r"^csv-\d{3}-[a-z0-9]+(?:-[a-z0-9]+)*$")


class ExportError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2)
    path.write_text(text + "\n", encoding="utf-8", newline="\n")


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        reader = csv.DictReader(handle, dialect=dialect)
        rows = [{key: (value or "") for key, value in row.items()} for row in reader]
        return list(reader.fieldnames or []), rows


def normalize_excel(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def verify_csv_xlsx_equal(csv_headers: list[str], csv_rows: list[dict[str, str]], xlsx_path: Path) -> None:
    workbook = load_workbook(xlsx_path, data_only=False, read_only=False)
    if workbook.sheetnames != ["Inventar"]:
        raise ExportError(f"XLSX sheets unexpected: {workbook.sheetnames!r}")
    sheet = workbook["Inventar"]
    expected_rows = len(csv_rows) + 1
    expected_cols = len(csv_headers)
    if sheet.max_row != expected_rows or sheet.max_column != expected_cols:
        raise ExportError(
            f"XLSX range mismatch: rows={sheet.max_row}, cols={sheet.max_column}; "
            f"expected rows={expected_rows}, cols={expected_cols}"
        )
    xlsx_headers = [normalize_excel(sheet.cell(1, col).value) for col in range(1, expected_cols + 1)]
    if xlsx_headers != csv_headers:
        raise ExportError("CSV/XLSX header mismatch")
    for row_index, csv_row in enumerate(csv_rows, start=2):
        for col_index, header in enumerate(csv_headers, start=1):
            xlsx_value = normalize_excel(sheet.cell(row_index, col_index).value)
            csv_value = csv_row[header]
            if xlsx_value != csv_value:
                raise ExportError(
                    f"CSV/XLSX mismatch at row {row_index}, column {header}: "
                    f"CSV={csv_value!r}, XLSX={xlsx_value!r}"
                )


def split_semicolon(value: str) -> list[str]:
    return [part.strip() for part in value.split(";") if part.strip()]


def split_pipe(value: str) -> list[str]:
    return [part.strip() for part in value.split("|") if part.strip()]


def to_int(value: str, field: str, product_id: str) -> int:
    try:
        return int(value)
    except ValueError as exc:
        raise ExportError(f"Invalid integer {field}={value!r} for {product_id}") from exc


def to_number(value: str, field: str, product_id: str) -> int | float:
    normalized = value.strip().replace(",", ".")
    try:
        number = float(normalized)
    except ValueError as exc:
        raise ExportError(f"Invalid number {field}={value!r} for {product_id}") from exc
    return int(number) if number.is_integer() else number


def to_bool_ja(value: str, field: str, product_id: str) -> bool:
    if value == "JA":
        return True
    if value == "NEIN":
        return False
    raise ExportError(f"Invalid boolean {field}={value!r} for {product_id}")


def require_hash(path: Path, expected: str, label: str) -> str:
    actual = sha256_file(path)
    if actual != expected:
        raise ExportError(f"Hash mismatch for {label}: expected {expected}, got {actual}")
    return actual


def assert_fields_present(headers: Iterable[str], required: Iterable[str], label: str) -> None:
    header_set = set(headers)
    missing = sorted(set(required) - header_set)
    if missing:
        raise ExportError(f"Missing fields in {label}: {missing}")


def nested_product(row: dict[str, str], images: dict[str, Any]) -> dict[str, Any]:
    product_id = row["ID"]
    return {
        "alkohol": {
            "prozent": to_number(row["Alkoholgehalt_Prozent"], "Alkoholgehalt_Prozent", product_id),
            "status": row["Alkoholisch_Status"],
        },
        "beschreibung": row["Beschreibung_sichtbar_FINAL"],
        "bestand": {
            "aktiv": to_bool_ja(row["Bestand_Aktiv"], "Bestand_Aktiv", product_id),
            "anzahl": to_int(row["Bestand_Anzahl"], "Bestand_Anzahl", product_id),
            "einheit": row["Bestand_Einheit"],
            "fuellstand": row["Füllstand"],
            "mengeText": row["Menge"],
            "oeffnungsstatus": row["Status"],
            "statusCode": row["Bestand_Status_Strukturiert"],
        },
        "bilder": images,
        "flasche": {
            "groesseMl": to_int(row["Flaschengroesse_ml"], "Flaschengroesse_ml", product_id),
        },
        "geschmacksTags": split_semicolon(row["FlavorTags_sichtbar"]),
        "herkunft": row["Herkunft_sichtbar"],
        "hersteller": row["Hersteller"],
        "id": product_id,
        "kategorie": row["Kategorie"],
        "logik": {
            "appFilterPrimaer": row["App_Filter_Primaer"],
            "appFilterSekundaer": row["App_Filter_Sekundaer"],
            "cocktailRelevanz": row["Cocktail_Relevanz_Status"],
            "cocktailrolle": row["Einsatz_Cocktailrolle"],
            "empfohleneCocktailtypen": split_pipe(row["Empfohlene_Cocktailtypen"]),
            "flavorPrimaer": row["Flavor_Profile_Primaer"],
            "flavorSekundaer": split_pipe(row["Flavor_Profile_Sekundaer"]),
            "gastmodusRelevanz": row["Guest_Mode_Relevanz"],
            "mixerGarnishSirupAbgrenzung": row["Mixer_Garnish_Sirup_Abgrenzung"],
            "produktart": row["Produktart_Strukturiert"],
        },
        "name": row["Name_final"],
        "nr": to_int(row["Nr"], "Nr", product_id),
        "nutzungsTags": split_semicolon(row["UsageTags_sichtbar"]),
        "unterkategorie": row["Unterkategorie"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--master-csv", required=True, type=Path)
    parser.add_argument("--master-xlsx", required=True, type=Path)
    parser.add_argument("--asset-package-zip", required=True, type=Path)
    parser.add_argument("--asset-mapping-csv", required=True, type=Path)
    parser.add_argument("--asset-manifest-json", required=True, type=Path)
    parser.add_argument("--asset-dir", required=True, type=Path)
    parser.add_argument("--phase0-matrix", required=True, type=Path)
    parser.add_argument("--phase0-audit-summary", required=True, type=Path)
    parser.add_argument("--manifest-v200", required=True, type=Path)
    parser.add_argument("--historical-app-zip", required=True, type=Path)
    parser.add_argument("--site", required=True, type=Path)
    args = parser.parse_args()

    try:
        verify_environment(verify_wheels=False, exact_python=False)
    except EnvironmentContractError as exc:
        raise ExportError(f"Environment contract failed: {exc}") from exc
    global load_workbook, Image
    from openpyxl import load_workbook as _load_workbook
    from PIL import Image as _Image
    load_workbook = _load_workbook
    Image = _Image

    lock = load_json(CONFIG_DIR / "source-lock.json")
    allowlist = load_json(CONFIG_DIR / "public-field-allowlist.json")
    version = load_json(CONFIG_DIR / "app-version.json")
    source_lock = lock["sources"]

    source_paths = {
        "masterCsv": args.master_csv,
        "masterXlsx": args.master_xlsx,
        "stableAssetPackageZip": args.asset_package_zip,
        "assetMappingCsv": args.asset_mapping_csv,
        "assetManifestJson": args.asset_manifest_json,
        "phase0TransferMatrix": args.phase0_matrix,
        "phase0AuditSummary": args.phase0_audit_summary,
        "manifestV200": args.manifest_v200,
        "historicalAppReadOnly": args.historical_app_zip,
    }
    source_hashes: dict[str, str] = {}
    for key, path in source_paths.items():
        if not path.is_file():
            raise ExportError(f"Missing source {key}: {path}")
        source_hashes[key] = require_hash(path, source_lock[key]["sha256"], key)

    master_headers, master_rows = read_csv(args.master_csv)
    if len(master_headers) != source_lock["masterCsv"]["columns"]:
        raise ExportError(f"Master column count mismatch: {len(master_headers)}")
    if len(master_rows) != source_lock["masterCsv"]["records"]:
        raise ExportError(f"Master row count mismatch: {len(master_rows)}")

    assert_fields_present(
        master_headers,
        allowlist["sourceFieldsAllowed"] + allowlist["sourceFieldsRequiredForValidationOnly"],
        "master CSV",
    )
    classified_master_fields = (
        set(allowlist["sourceFieldsAllowed"])
        | set(allowlist["sourceFieldsRequiredForValidationOnly"])
        | set(allowlist["sourceFieldsForbiddenInPublicBuild"])
    )
    if set(master_headers) != classified_master_fields:
        raise ExportError(
            f"Master field classification incomplete: "
            f"unclassified={sorted(set(master_headers)-classified_master_fields)}, "
            f"unknown={sorted(classified_master_fields-set(master_headers))}"
        )
    verify_csv_xlsx_equal(master_headers, master_rows, args.master_xlsx)

    ids: set[str] = set()
    numbers: set[int] = set()
    names: set[str] = set()
    for row in master_rows:
        product_id = row["ID"]
        if not ID_RE.fullmatch(product_id):
            raise ExportError(f"Invalid stable product ID: {product_id!r}")
        if product_id in ids:
            raise ExportError(f"Duplicate product ID: {product_id}")
        ids.add(product_id)
        number = to_int(row["Nr"], "Nr", product_id)
        if number in numbers:
            raise ExportError(f"Duplicate visible number: {number}")
        numbers.add(number)
        if row["Name_final"] in names:
            raise ExportError(f"Duplicate final name: {row['Name_final']}")
        names.add(row["Name_final"])
        if row["OK_FINAL_MASTER"] != "OK_FINAL_MASTER":
            raise ExportError(f"Product not final: {product_id}")
        if row["Beschreibung_Status"] != "OK_FINAL_MASTER":
            raise ExportError(f"Visible description not final: {product_id}")
        if row["Datenqualitaet_App_Logik"] != "APP_READY":
            raise ExportError(f"Product not APP_READY: {product_id}")
        allowed_mapping_statuses = {
            "DATEIGENAU_BESTAETIGT",
            "DATEIGENAU_BESTAETIGT_NACH_USER_KLAERUNG",
            "DATEIGENAU_BESTAETIGBAR_FUER_SAMMELPOSITION",
        }
        if row["Bildmapping_Status"] not in allowed_mapping_statuses:
            raise ExportError(f"Image mapping not confirmed: {product_id}")

    mapping_headers, mapping_rows = read_csv(args.asset_mapping_csv)
    if len(mapping_headers) != source_lock["assetMappingCsv"]["columns"]:
        raise ExportError(f"Asset mapping column count mismatch: {len(mapping_headers)}")
    if len(mapping_rows) != source_lock["assetMappingCsv"]["records"]:
        raise ExportError(f"Asset mapping row count mismatch: {len(mapping_rows)}")
    assert_fields_present(
        mapping_headers,
        allowlist["assetFieldsAllowed"] + allowlist["assetFieldsRequiredForValidationOnly"],
        "asset mapping CSV",
    )
    classified_asset_fields = (
        set(allowlist["assetFieldsAllowed"])
        | set(allowlist["assetFieldsRequiredForValidationOnly"])
        | set(allowlist["assetFieldsForbiddenInPublicBuild"])
    )
    if set(mapping_headers) != classified_asset_fields:
        raise ExportError(
            f"Asset field classification incomplete: "
            f"unclassified={sorted(set(mapping_headers)-classified_asset_fields)}, "
            f"unknown={sorted(classified_asset_fields-set(mapping_headers))}"
        )

    master_by_id = {row["ID"]: row for row in master_rows}
    by_product: dict[str, dict[str, Any]] = defaultdict(lambda: {"haupt": None, "zusatz": []})
    public_assets: list[dict[str, Any]] = []
    asset_paths: set[str] = set()
    primary_count = 0
    additional_count = 0

    for row in mapping_rows:
        product_id = row["ID"]
        if product_id not in ids:
            raise ExportError(f"Asset references unknown product: {product_id}")
        master_row = master_by_id[product_id]
        if row["Nr_sichtbar"] != master_row["Nr"]:
            raise ExportError(f"Asset visible number mismatch: {product_id}")
        if row["Name_final"] != master_row["Name_final"]:
            raise ExportError(f"Asset product name mismatch: {product_id}")
        if row["Hersteller"] != master_row["Hersteller"]:
            raise ExportError(f"Asset manufacturer mismatch: {product_id}")
        if row["Zuordnungsstatus"] not in {"BESTAETIGT", "BESTAETIGT_SONDERFALL"}:
            raise ExportError(f"Asset mapping status not confirmed: {product_id}")
        if row["Nutzerbestaetigt"] != "JA":
            raise ExportError(f"Asset mapping not user-confirmed: {product_id}")
        if row["Dateiformat"] != "JPEG":
            raise ExportError(f"Unexpected public asset format: {product_id}")
        role_raw = row["Bildrolle"]
        if role_raw == "HAUPTBILD":
            role = "haupt"
            primary_count += 1
        elif role_raw == "ZUSATZBILD":
            role = "zusatz"
            additional_count += 1
        else:
            raise ExportError(f"Unexpected asset role {role_raw!r} for {product_id}")

        relative_path = row["Relativer_Assetpfad"].replace("\\", "/")
        if relative_path.startswith("/") or ".." in Path(relative_path).parts:
            raise ExportError(f"Unsafe asset path: {relative_path}")
        if relative_path in asset_paths:
            raise ExportError(f"Duplicate asset path: {relative_path}")
        asset_paths.add(relative_path)
        filename = Path(relative_path).name
        if row["Kanonischer_Dateiname"] != filename:
            raise ExportError(f"Canonical asset filename mismatch: {relative_path}")
        if not filename.startswith(product_id + "__"):
            raise ExportError(f"Asset filename is not ID-based: {filename} for {product_id}")

        source_asset = args.asset_dir / relative_path
        if not source_asset.is_file():
            raise ExportError(f"Missing stable asset: {source_asset}")
        actual_hash = sha256_file(source_asset)
        if actual_hash != row["Asset_SHA256"]:
            raise ExportError(f"Asset hash mismatch: {relative_path}")
        actual_size = source_asset.stat().st_size
        if actual_size != to_int(row["Asset_Dateigroesse_Byte"], "Asset_Dateigroesse_Byte", product_id):
            raise ExportError(f"Asset size mismatch: {relative_path}")
        with Image.open(source_asset) as image:
            width, height = image.size
            if width != to_int(row["Breite_Pixel"], "Breite_Pixel", product_id):
                raise ExportError(f"Asset width mismatch: {relative_path}")
            if height != to_int(row["Hoehe_Pixel"], "Hoehe_Pixel", product_id):
                raise ExportError(f"Asset height mismatch: {relative_path}")
            if image.format != "JPEG":
                raise ExportError(f"Asset is not JPEG: {relative_path}")
        expected_orientation = "HOCHFORMAT" if height > width else "QUERFORMAT" if width > height else "QUADRAT"
        if row["Ausrichtung"] != expected_orientation:
            raise ExportError(f"Asset orientation mismatch: {relative_path}")

        public_asset = {
            "breitePixel": width,
            "dateigroesseByte": actual_size,
            "hoehePixel": height,
            "id": product_id,
            "mimeTyp": "image/jpeg",
            "pfad": relative_path,
            "rolle": role,
            "sha256": actual_hash,
        }
        public_assets.append(public_asset)
        if role == "haupt":
            if by_product[product_id]["haupt"] is not None:
                raise ExportError(f"Multiple primary assets for {product_id}")
            by_product[product_id]["haupt"] = relative_path
        else:
            by_product[product_id]["zusatz"].append(relative_path)

    if primary_count != 142 or additional_count != 12:
        raise ExportError(f"Unexpected asset roles: primary={primary_count}, additional={additional_count}")
    for product_id in sorted(ids):
        if by_product[product_id]["haupt"] is None:
            raise ExportError(f"Missing primary asset for {product_id}")
        by_product[product_id]["zusatz"].sort()

    products = [nested_product(row, by_product[row["ID"]]) for row in master_rows]
    products.sort(key=lambda item: item["nr"])
    public_assets.sort(key=lambda item: (item["id"], 0 if item["rolle"] == "haupt" else 1, item["pfad"]))

    site = args.site.resolve()
    data_dir = site / "data"
    target_assets_dir = site / "assets" / "images" / "inventory"
    if target_assets_dir.exists():
        shutil.rmtree(target_assets_dir)
    target_assets_dir.mkdir(parents=True, exist_ok=True)

    for public_asset in public_assets:
        source_asset = args.asset_dir / public_asset["pfad"]
        target_asset = site / public_asset["pfad"]
        target_asset.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source_asset, target_asset)
        if sha256_file(target_asset) != public_asset["sha256"]:
            raise ExportError(f"Copied asset hash mismatch: {public_asset['pfad']}")

    inventory_export = {
        "items": products,
        "masterVersion": version["masterVersion"],
        "schemaVersion": 1,
    }
    asset_export = {
        "assetVersion": version["assetVersion"],
        "items": public_assets,
        "schemaVersion": 1,
    }
    write_json(data_dir / "inventory.json", inventory_export)
    write_json(data_dir / "assets.json", asset_export)

    export_metadata = {
        "assetVersion": version["assetVersion"],
        "counts": {
            "additionalAssets": additional_count,
            "assetRecords": len(public_assets),
            "inventoryRecords": len(products),
            "primaryAssets": primary_count,
        },
        "historicalAppUsedAsInput": False,
        "masterVersion": version["masterVersion"],
        "policy": {
            "allowlistSha256": sha256_file(CONFIG_DIR / "public-field-allowlist.json"),
            "mode": allowlist["policy"],
            "privateFieldsSerialized": 0,
        },
        "schemaVersion": 1,
        "sourceHashes": source_hashes,
    }
    write_json(data_dir / "export-metadata.json", export_metadata)

    print(json.dumps({
        "status": "PASS",
        "inventoryRecords": len(products),
        "assetRecords": len(public_assets),
        "site": str(site),
    }, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ExportError as exc:
        print(f"EXPORT_ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2)
